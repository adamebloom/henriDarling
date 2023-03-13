import openpyxl
import unidecode
import export
import os


def scrap_data(file):
    wb = openpyxl.load_workbook(file, read_only=True)
    sheets = wb.worksheets
    to_export = []
    for sheet in sheets:
        if sheet['C6'].value == 'Inactif' or str(sheet['C6'].value).startswith('Sit.'):
            continue
        else:
            data = []
            name = ['Name', sheet['B1'].value]
            data.append(name)
            direction_row = -1
            for row in sheet.iter_rows(max_row=17, max_col=4, min_row=13, min_col=3, values_only=True):
                data.append(list(row))

            for row in sheet.iter_rows(max_col=2, min_col=1, min_row=30, values_only=False):
                if str(row[0].value).startswith('Effectif'):
                    data.append(list((row[0].value, row[1].value)))
                if str(row[0].value).startswith('Direction'):
                    direction_row = row[0].row

            if direction_row > -1:
                list_direction = []
                for row in sheet.iter_rows(min_col=3, min_row=direction_row, values_only=True):
                    if row[0]:
                        directions = [list(xi for xi in x if xi is not None) for x in row if x is not None]
                        new_list = [''.join(sous_list) for sous_list in directions]
                        cleaned_list = [unidecode.unidecode(word) for word in new_list]
                        cleaned_list.pop(0)
                        final = ''
                        for word in cleaned_list:
                            final += ''.join(word) + ' '
                        list_direction.append(final)
                data.append(list(("Direction", list_direction)))

            data = list(xi for xi in data if xi[0] is not None)
            nope = ['Fax', 'Langue']
            data = list(xi for xi in data if xi[0] not in nope)

            for element in data:
                if element[0].startswith('Téléphone'):
                    element[0] = 'Phone'
                if element[0].startswith('E-mail'):
                    element[0] = 'Mail'
                if element[0].startswith('Site internet'):
                    element[0] = 'Website'
                if element[0].startswith('Effectif'):
                    element[0] = 'Effectifs'
            to_export.append(data)
    wb.close()
    return to_export


if __name__ == "__main__":
    directory = "./convert"
    infos = []
    for filename in os.listdir(directory):
        f = os.path.join(directory, filename)
        if os.path.isfile(f):
            infos += scrap_data(f)
    export.to_excel(infos)
